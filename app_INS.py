import streamlit as st
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.naive_bayes import MultinomialNB
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, GridSearchCV, StratifiedKFold
from sklearn.metrics import accuracy_score, classification_report
from sklearn.preprocessing import LabelEncoder, OneHotEncoder
from sklearn.pipeline import Pipeline
from imblearn.over_sampling import SMOTE
from imblearn.pipeline import Pipeline as ImbPipeline
import plotly.graph_objects as go
import re
from collections import Counter
import os
import pickle
import hashlib
from datetime import datetime
import base64
import time
from io import BytesIO
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


def get_base64_of_bin_file(bin_file):
    """Convertit un fichier binaire en base64"""
    try:
        with open(bin_file, 'rb') as f:
            data = f.read()
        return base64.b64encode(data).decode()
    except:
        return None

def set_background_image(image_file):
    """Définit une image de fond avec overlay noir glacial"""
    bin_str = get_base64_of_bin_file(image_file)
    if bin_str:
        page_bg_img = f'''
        <style>
        .stApp {{
            background-image: url("data:image/png;base64,{bin_str}");
            background-size: cover;
            background-position: center;
            background-repeat: no-repeat;
            background-attachment: fixed;
        }}
        .stApp::before {{
            content: "";
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: linear-gradient(135deg, 
                rgba(0, 0, 0, 0.85) 0%, 
                rgba(10, 25, 40, 0.9) 25%, 
                rgba(0, 15, 30, 0.88) 50%, 
                rgba(5, 20, 35, 0.9) 75%, 
                rgba(0, 0, 0, 0.85) 100%);
            backdrop-filter: blur(1px);
            z-index: -1;
        }}
        </style>
        '''
        st.markdown(page_bg_img, unsafe_allow_html=True)




# Configuration de la page
st.set_page_config(
    page_title="Classifieur de l’activité selon la nomenclature",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Définir l'image de fond
image_path = os.path.join("images", "Structure.jpg")
#image_path = os.path.abspath(os.path.join("images", "Structure.jpg"))
if os.path.exists(image_path):
    set_background_image(image_path)

# Afficher les logos
col1, col2, col3 = st.columns([1, 4, 1])
with col1:
    logo1_path = os.path.join("images", "logo1.jpg")
    if os.path.exists(logo1_path):
        st.image(logo1_path, use_column_width=True)

with col3:
    logo2_path = os.path.join("images", "logo2.jpg")
    if os.path.exists(logo2_path):
        st.image(logo2_path, use_column_width=True)


def display_fixed_logos():
    """Affiche les logos fixes - logo1 à gauche, logo2 à droite - version ajustée"""
    logo1_str = get_base64_of_bin_file("logo1.png")
    logo2_str = get_base64_of_bin_file("logo2.png")
    
    if not logo1_str and not logo2_str:
        st.sidebar.info("Logos non trouvés. Placez logo1.png et logo2.png dans le répertoire de l'application.")
        return
    
    # CSS avec position ajustée pour éviter la barre Streamlit et taille augmentée
    logos_css = """
    <style>
    .fixed-logo-left {
        position: fixed !important;
        top: 80px !important;
        left: 20px !important;
        z-index: 9999 !important;
        pointer-events: none !important;
    }
    
    .fixed-logo-right {
        position: fixed !important;
        top: 80px !important;
        right: 20px !important;
        z-index: 9999 !important;
        pointer-events: none !important;
    }
    
    .fixed-logo-left img, .fixed-logo-right img {
        height: 200px !important;
        width: auto !important;
        border-radius: 12px !important;
        box-shadow: 0 6px 20px rgba(0,0,0,0.4) !important;
        transition: all 0.3s ease !important;
        background: rgba(255,255,255,0.15) !important;
        padding: 8px !important;
        backdrop-filter: blur(10px) !important;
        border: 2px solid rgba(255,255,255,0.2) !important;
        display: block !important;
        pointer-events: auto !important;
    }
    
    .fixed-logo-left img:hover, .fixed-logo-right img:hover {
        transform: scale(1.05) !important;
        box-shadow: 0 8px 25px rgba(0,0,0,0.5) !important;
        background: rgba(255,255,255,0.2) !important;
    }
    
    /* Assurer la visibilité sur tous les thèmes */
    .fixed-logo-left, .fixed-logo-right {
        background: transparent !important;
    }
    
    /* Adaptation responsive */
    @media (max-width: 768px) {
        .fixed-logo-left img, .fixed-logo-right img {
            height: 80px !important;
        }
        .fixed-logo-left {
            left: 10px !important;
        }
        .fixed-logo-right {
            right: 10px !important;
        }
    }
    </style>
    """
    
    st.markdown(logos_css, unsafe_allow_html=True)
    
    # Afficher les logos séparément pour un meilleur contrôle
    if logo1_str:
        logo1_html = f'''
        <div class="fixed-logo-left">
            <img src="data:image/png;base64,{logo1_str}" alt="Logo 1" title="Institut National de la Statistique">
        </div>
        '''
        st.markdown(logo1_html, unsafe_allow_html=True)
    
    if logo2_str:
        logo2_html = f'''
        <div class="fixed-logo-right">
            <img src="data:image/png;base64,{logo2_str}" alt="Logo 2" title="Institut Sous-régional de Statistique et d'Économie Appliquée">
        </div>
        '''
        st.markdown(logo2_html, unsafe_allow_html=True)
    
    # Message de confirmation dans la sidebar

# Déplacer de manière transitoire l'arrière plan
st.markdown("""
<style>
/* Décalage du contenu principal */
[data-testid="stSidebar"][aria-expanded="true"] ~ div[data-testid="stAppViewContainer"] {
    margin-left: 250px;
    transition: margin-left 0.3s ease;
}
[data-testid="stSidebar"][aria-expanded="false"] ~ div[data-testid="stAppViewContainer"] {
    margin-left: 0;
    transition: margin-left 0.3s ease;
}

/* ✅ Correction : déplacement direct du logo gauche */
.fixed-logo-left {
    position: fixed !important;
    top: 80px !important;
    left: 20px !important;
    transition: left 0.3s ease !important;
    z-index: 9999 !important;
}

/* Quand la sidebar est ouverte (hack via parent [aria-expanded]) */
section[data-testid="stSidebar"][aria-expanded="true"] ~ div [class="fixed-logo-left"] {
    left: 270px !important; /* 250px + marge */
}

section[data-testid="stSidebar"][aria-expanded="true"] ~ div [class="fixed-logo-left"] {
    left: 270px !important;
    opacity: 0.85;
    transition: left 0.3s ease, opacity 0.3s ease;
}

</style>
""", unsafe_allow_html=True)

# Barre supérieure
st.markdown("""
<style>
header[data-testid="stHeader"] {
    background-color: #567671 !important;
}
</style>
""", unsafe_allow_html=True)

# Barre de défilement
st.markdown("""
<style>
/* --- Sidebar Scrollbar --- */
section[data-testid="stSidebar"]::-webkit-scrollbar {
    width: 40px;
}
section[data-testid="stSidebar"]::-webkit-scrollbar-track {
    background: #08C478;
    border-radius: 40px;
}
section[data-testid="stSidebar"]::-webkit-scrollbar-thumb {
    background: linear-gradient(180deg, #4B0082, #8A2BE2);
    border-radius: 40px;
}
section[data-testid="stSidebar"]::-webkit-scrollbar-thumb:hover {
    background: linear-gradient(180deg, #5E11A3, #A15CF2);
}

/* --- Contenu principal Scrollbar --- */
div[data-testid="stAppViewContainer"]::-webkit-scrollbar {
    width: 40px;
}
div[data-testid="stAppViewContainer"]::-webkit-scrollbar-thumb {
    background-color: #444;
    border-radius: 40px;
}
div[data-testid="stAppViewContainer"]::-webkit-scrollbar-thumb:hover {
    background-color: #777;
}
</style>
""", unsafe_allow_html=True)


def load_css():
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');
        
        * {
            font-family: 'Poppins', sans-serif;
        }
        
        .main-header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 3rem 2rem;
            border-radius: 20px;
            margin-bottom: 2rem;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            animation: slideInDown 0.8s ease-out;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.2);
        }
        
        .main-header h1 {
            color: white;
            text-align: center;
            margin: 0;
            font-size: 3rem;
            font-weight: 700;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }
        
        .main-header p {
            color: rgba(255,255,255,0.9);
            text-align: center;
            font-size: 1.3rem;
            margin: 1rem 0 0 0;
        }
        
        /* Amélioration de la lisibilité de la sidebar */
        .css-1d391kg {
            background-color: rgba(248, 250, 252, 0.98) !important;
        }
        
        .stSidebar {
            background-color: rgba(248, 250, 252, 0.98) !important;
        }
        
        .stSidebar .stMarkdown {
            color: #1a202c !important;
        }
        
        .stSidebar h1, .stSidebar h2, .stSidebar h3, .stSidebar h4, .stSidebar h5, .stSidebar h6 {
            color: #2d3748 !important;
            text-shadow: none !important;
            font-weight: 600 !important;
        }
        
        .stSidebar p, .stSidebar span, .stSidebar div, .stSidebar label {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        /* Amélioration des selectbox */
        .stSidebar .stSelectbox > div > div {
            background: #ffffff !important;
            color: #1a202c !important;
            border: 2px solid #cbd5e0 !important;
            border-radius: 8px !important;
            font-weight: 500 !important;
        }
        
        .stSidebar .stSelectbox > div > div > div {
            color: #1a202c !important;
            font-weight: 500 !important;
        }
        
        .stSidebar .stSelectbox > div > div:hover {
            border-color: #667eea !important;
            box-shadow: 0 0 0 1px #667eea !important;
        }
        
        /* Amélioration des checkbox */
        .stSidebar .stCheckbox > label {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        .stSidebar .stCheckbox > label > span {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        /* Amélioration des boutons dans la sidebar */
        .stSidebar .stButton > button {
            background: #667eea !important;
            color: #ffffff !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            transition: all 0.2s ease !important;
        }
        
        .stSidebar .stButton > button:hover {
            background: #5a67d8 !important;
            transform: translateY(-1px) !important;
        }
        
        /* Amélioration des file uploader */
        .stSidebar .stFileUploader > div {
            background: #ffffff !important;
            border: 2px dashed #cbd5e0 !important;
            border-radius: 8px !important;
        }
        
        .stSidebar .stFileUploader label {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        /* Amélioration des alertes dans la sidebar */
        .stSidebar .stAlert {
            background: #ffffff !important;
            color: #2d3748 !important;
            border-radius: 8px !important;
            border-left: 4px solid #667eea !important;
        }
        
        .stSidebar .stSuccess {
            background: #f0fff4 !important;
            color: #22543d !important;
            border-left-color: #38a169 !important;
        }
        
        .stSidebar .stWarning {
            background: #fffbf0 !important;
            color: #744210 !important;
            border-left-color: #d69e2e !important;
        }
        
        .stSidebar .stError {
            background: #fff5f5 !important;
            color: #742a2a !important;
            border-left-color: #e53e3e !important;
        }
        
        .stSidebar .stInfo {
            background: #ebf8ff !important;
            color: #2a4365 !important;
            border-left-color: #3182ce !important;
        }
        
        /* Amélioration des expandeurs */
        .stSidebar .stExpander {
            background: #ffffff !important;
            border: 1px solid #e2e8f0 !important;
            border-radius: 8px !important;
        }
        
        .stSidebar .stExpander > div > div > div {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        /* Amélioration du text input et text area */
        .stSidebar .stTextInput > div > div > input {
            background: #ffffff !important;
            color: #1a202c !important;
            border: 2px solid #cbd5e0 !important;
            border-radius: 8px !important;
        }
        
        .stSidebar .stTextArea > div > div > textarea {
            background: #ffffff !important;
            color: #1a202c !important;
            border: 2px solid #cbd5e0 !important;
            border-radius: 8px !important;
        }
        
        /* Amélioration globale des éléments de la sidebar */
        .stSidebar .element-container {
            color: #2d3748 !important;
        }
        
        .stSidebar .markdown-text-container {
            color: #2d3748 !important;
        }
        
        .level-card {
            background: rgba(15, 25, 35, 0.95);
            color: #ffffff;
            padding: 2rem;
            border-radius: 15px;
            border-left: 5px solid #007bff;
            margin: 1.5rem 0;
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
            transition: all 0.3s ease;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.1);
        }
        
        .prediction-result {
            background: rgba(15, 35, 25, 0.95);
            color: #ffffff;
            padding: 2rem;
            border-radius: 15px;
            border: 3px solid #28a745;
            margin: 1.5rem 0;
            backdrop-filter: blur(10px);
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
        }
        
        .hierarchy-selector {
            background: rgba(255, 255, 255, 0.9);
            color: #1f2937;
            padding: 1.5rem;
            border-radius: 15px;
            border: 2px solid rgba(107, 114, 128, 0.2);
            margin: 1rem 0;
            backdrop-filter: blur(10px);
        }
        
        .update-section {
            background: rgba(15, 35, 25, 0.95);
            color: #ffffff;
            padding: 1.5rem;
            border-radius: 15px;
            border: 2px solid #28a745;
            margin: 1rem 0;
            backdrop-filter: blur(10px);
        }
        
        .batch-section {
            background: rgba(25, 15, 35, 0.95);
            color: #ffffff;
            padding: 1.5rem;
            border-radius: 15px;
            border: 2px solid #764ba2;
            margin: 1rem 0;
            backdrop-filter: blur(10px);
        }
        
        .hierarchy-info {
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.3) 0%, rgba(118, 75, 162, 0.3) 100%);
            padding: 1rem;
            border-radius: 10px;
            border: 1px solid rgba(255,255,255,0.2);
            margin: 0.5rem 0;
        }
        
        /* Amélioration des selectbox globales */
        .stSelectbox > div > div {
            background: #ffffff !important;
            color: #1a202c !important;
            border: 2px solid #cbd5e0 !important;
            border-radius: 8px !important;
            font-weight: 500 !important;
        }
        
        .stSelectbox > div > div > div {
            color: #1a202c !important;
            font-weight: 500 !important;
        }
        
        .stSelectbox > div > div:hover {
            border-color: #667eea !important;
            box-shadow: 0 0 0 1px #667eea !important;
        }
        
        /* Amélioration des text areas globales */
        .stTextArea > div > div > textarea {
            background: #ffffff !important;
            color: #1a202c !important;
            border: 2px solid #cbd5e0 !important;
            border-radius: 8px !important;
            font-weight: 500 !important;
        }
        
        .stTextArea > div > div > textarea::placeholder {
            color: #718096 !important;
            font-weight: 400 !important;
        }
        
        .stTextArea > div > div > textarea:focus {
            border-color: #667eea !important;
            box-shadow: 0 0 0 1px #667eea !important;
        }
        
        /* Amélioration des dataframes */
        .stDataFrame {
            background: #ffffff !important;
            border-radius: 8px !important;
            overflow: hidden !important;
        }
        
        .stDataFrame table {
            color: #1a202c !important;
        }
        
        .stDataFrame th {
            background: #f7fafc !important;
            color: #2d3748 !important;
            font-weight: 600 !important;
        }
        
        .stDataFrame td {
            color: #2d3748 !important;
        }
        
        /* Amélioration des métriques et informations */
        .metric-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 2rem;
            border-radius: 15px;
            text-align: center;
            margin: 1rem 0;
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
        }
        
        .hierarchy-info {
            background: linear-gradient(135deg, rgba(102, 126, 234, 0.1) 0%, rgba(118, 75, 162, 0.1) 100%);
            color: #1a202c;
            padding: 1rem;
            border-radius: 10px;
            border: 2px solid rgba(102, 126, 234, 0.2);
            margin: 0.5rem 0;
        }
        
        /* Amélioration des sections colorées */
        .hierarchy-selector {
            background: rgba(255, 255, 255, 0.95);
            color: #1a202c;
            padding: 1.5rem;
            border-radius: 15px;
            border: 2px solid rgba(102, 126, 234, 0.2);
            margin: 1rem 0;
            backdrop-filter: blur(10px);
        }
        
        .update-section {
            background: rgba(240, 253, 244, 0.95);
            color: #1a202c;
            padding: 1.5rem;
            border-radius: 15px;
            border: 2px solid #68d391;
            margin: 1rem 0;
            backdrop-filter: blur(10px);
        }
        
        .batch-section {
            background: rgba(252, 245, 255, 0.95);
            color: #1a202c;
            padding: 1.5rem;
            border-radius: 15px;
            border: 2px solid #b794f6;
            margin: 1rem 0;
            backdrop-filter: blur(10px);
        }
        
        /* Amélioration des file uploaders globaux */
        .stFileUploader > div {
            background: #ffffff !important;
            border: 2px dashed #cbd5e0 !important;
            border-radius: 8px !important;
        }
        
        .stFileUploader label {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        /* Amélioration des sliders */
        .stSlider > div > div > div > div {
            background: #667eea !important;
        }
        
        .stSlider > div > div > div > div > div {
            background: #667eea !important;
        }
        
        /* Amélioration des expander globaux */
        .stExpander {
            background: #ffffff !important;
            border: 1px solid #e2e8f0 !important;
            border-radius: 8px !important;
        }
        
        .stExpander > div > div > div {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        /* Amélioration des colonnes et containers */
        .element-container {
            color: #2d3748 !important;
        }
        
        .markdown-text-container {
            color: #2d3748 !important;
        }
        
        /* Amélioration du contraste pour les labels */
        label {
            color: #2d3748 !important;
            font-weight: 500 !important;
        }
        
        .confidence-high { color: #28a745; font-weight: bold; }
        .confidence-medium { color: #ffc107; font-weight: bold; }
        .confidence-low { color: #dc3545; font-weight: bold; }
        
        h1, h2, h3, h4, h5, h6 { color: #ffffff !important; text-shadow: 2px 2px 4px rgba(0,0,0,0.5); }
        .stMarkdown { color: #ffffff; }
        p, span, div { color: #ffffff !important; }

        /* Info-bulles et tooltips : fond sombre, texte blanc pour meilleur contraste (sans bordures) */
        [data-baseweb="tooltip"],
        [data-baseweb="popover"],
        .stTooltipIcon > span,
        div[data-testid="stToolbar"] span {
            color: #ffffff !important;
            font-size: 1.02rem !important;
            font-weight: 600 !important;
            line-height: 1.4 !important;
            background: rgba(10, 10, 12, 0.90) !important;
            padding: 0.6rem 0.9rem !important;
            border-radius: 8px !important;
            border: 0 !important;
            outline: none !important;
            box-shadow: 0 8px 24px rgba(0, 0, 0, 0.6) !important;
            max-width: 360px !important;
            backdrop-filter: blur(6px) !important;
            text-shadow: none !important;
            margin: 4px !important;
        }

        /* Style spécifique pour les tooltips Streamlit (contenu) - sans bordure */
        .stTooltipContent {
            background-color: rgba(10, 10, 12, 0.90) !important;
            color: #ffffff !important;
            border: 0 !important;
            outline: none !important;
            padding: 0.6rem !important;
            box-shadow: 0 8px 24px rgba(0, 0, 0, 0.6) !important;
            max-width: 360px !important;
        }

        /* Icône d'info : blanche pour contraste */
        [data-testid="stTooltipIcon"] > svg {
            fill: #ffffff !important;
            width: 20px !important;
            height: 20px !important;
            border: none !important;
        }

        .streamlit-expanderHeader:hover [data-testid="stTooltipIcon"] > svg {
            fill: #ffffff !important;
        }
        
        .stButton > button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 0.75rem 2rem;
            border-radius: 25px;
            font-weight: 600;
            transition: all 0.3s ease;
            box-shadow: 0 5px 15px rgba(0,0,0,0.2);
        }
        
        .stButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 25px rgba(0,0,0,0.3);
        }
    </style>
    """, unsafe_allow_html=True)

st.markdown("""
    <style>
    /* ✅ Boîte de texte plus lisible */
    textarea, .stTextArea textarea {
        background-color: #ffffff !important;
        color: #000000 !important;
        font-size: 16px !important;
    }

    /* ✅ Boutons mieux visibles */
    .stButton>button {
        background-color: #6c63ff !important;
        color: white !important;
        border-radius: 10px !important;
        font-weight: bold;
    }
    </style>
""", unsafe_allow_html=True)

class EnhancedHierarchicalPredictor:
    def __init__(self):
        self.vectorizer = TfidfVectorizer(
            max_features=10000,
            ngram_range=(1, 3),
            lowercase=True,
            analyzer='word',
            min_df=1,
            max_df=0.95
        )
        
        # Modèle unique pour CPU - MultinomialNB seulement
        self.base_models = {
            'Naive Bayes': MultinomialNB()#,
#            'Logistic Regression': LogisticRegression(max_iter=1000, solver='lbfgs', multi_class='auto')
        }
        
        # Grille de paramètres simplifiée
        self.param_grids = {
            'Naive Bayes': {'alpha': [0.1, 0.5, 1.0]}#,
#                 'Logistic Regression': {'C': [0.1, 1.0, 5.0], 'class_weight': [None, 'balanced']
#            }
        }
        
        # Structures de suivi
        self.label_encoders = {}
        self.hierarchy_predictors = {}
        self.best_models = {}
        self.hierarchy_structure = {}
        self.models_directory = "saved_models"
        self.use_smote = True
        
        if not os.path.exists(self.models_directory):
            os.makedirs(self.models_directory)
    
    # ----------------------------------------------------------------
    # UTILS
    # ----------------------------------------------------------------

    def get_data_hash(self, df):
        """Génère un hash unique pour les données"""
        data_string = pd.util.hash_pandas_object(df[['reponse', 'Classe', 'Grand poste', 'Section', 'Groupe']], index=True).values
        return hashlib.md5(str(data_string).encode()).hexdigest()[:8]
    
    def get_model_filename(self, data_hash, prediction_level):
        """Génère le nom de fichier pour sauvegarder le modèle"""
        return os.path.join(self.models_directory, f"enhanced_predictor_{prediction_level}_{data_hash}.pkl")
    
    def save_models(self, filename):
        """Sauvegarde tous les modèles et encodeurs optimisés"""
        model_data = {
            'vectorizer': self.vectorizer,
            'label_encoders': self.label_encoders,
            'hierarchy_predictors': self.hierarchy_predictors,
            'best_models': self.best_models,
            'hierarchy_structure': self.hierarchy_structure,
            'timestamp': datetime.now().isoformat(),
            'use_smote': self.use_smote
        }
        
        with open(filename, 'wb') as f:
            pickle.dump(model_data, f)
        
        st.success(f"Meilleurs modèles sauvegardés: {filename}")
    
    def load_models(self, filename):
        """Charge les modèles sauvegardés"""
        try:
            with open(filename, 'rb') as f:
                model_data = pickle.load(f)
            
            self.vectorizer = model_data['vectorizer']
            self.label_encoders = model_data['label_encoders']
            self.hierarchy_predictors = model_data.get('hierarchy_predictors', {})
            self.best_models = model_data.get('best_models', {})
            self.hierarchy_structure = model_data.get('hierarchy_structure', {})
            self.use_smote = model_data.get('use_smote', True)
            
            timestamp = model_data.get('timestamp', 'Inconnu')
            st.success(f"Modèles optimisés chargés (sauvegardés le: {timestamp[:19]})")
            return True
        except Exception as e:
            st.warning(f"Impossible de charger les modèles: {e}")
            return False
    
    def models_exist(self, filename):
        """Vérifie si les modèles sauvegardés existent"""
        return os.path.exists(filename)
    
    # ----------------------------------------------------------------
    # STRUCTURE HIÉRARCHIQUE
    # ----------------------------------------------------------------

    def build_hierarchy_structure(self, df):
        """Construit la structure hiérarchique complète"""
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        self.hierarchy_structure = {}
        
        for _, row in df.iterrows():
            current_level = self.hierarchy_structure
            
            for level in hierarchy_levels:
                value = row[level]
                if value not in current_level:
                    current_level[value] = {}
                current_level = current_level[value]
        
        return self.hierarchy_structure
    
    def get_filtered_options(self, level, parent_selections):
        """Obtient les options filtrées pour un niveau donné basé sur les sélections parentes"""
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        level_index = hierarchy_levels.index(level)
        
        current_structure = self.hierarchy_structure
        
        # Naviguer dans la structure selon les sélections parentes
        for i in range(level_index):
            parent_level = hierarchy_levels[i]
            if parent_level in parent_selections and parent_selections[parent_level]:
                if parent_selections[parent_level] in current_structure:
                    current_structure = current_structure[parent_selections[parent_level]]
                else:
                    return []
        
        return list(current_structure.keys()) if isinstance(current_structure, dict) else []
    
    def get_unique_prediction_path(self, selections):
        """Vérifie si le chemin sélectionné mène à une classe unique"""
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        current_structure = self.hierarchy_structure
        
        for level in hierarchy_levels:
            if level in selections and selections[level]:
                if selections[level] in current_structure:
                    current_structure = current_structure[selections[level]]
                    
                    # Si nous sommes au niveau Classe et qu'il n'y a qu'une seule option
                    if level == 'Classe' and len(current_structure) == 0:
                        return True, selections[level]
                    elif level != 'Classe' and isinstance(current_structure, dict) and len(current_structure) == 1:
                        # S'il n'y a qu'une seule option au niveau suivant
                        next_level_index = hierarchy_levels.index(level) + 1
                        if next_level_index < len(hierarchy_levels):
                            next_level = hierarchy_levels[next_level_index]
                            return False, list(current_structure.keys())[0]
                else:
                    return False, None
        
        return False, None
    
    # ----------------------------------------------------------------
    # PRÉTRAITEMENT
    # ----------------------------------------------------------------
    def preprocess_text(self, text, language='auto'):
        """Preprocessing du texte adapté à la langue"""
        text = str(text).lower()
        text = re.sub(r'[^\w\s\àâäéèêëïîôöùûüÿñç]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    
    def check_class_imbalance(self, y, threshold=0.1):
        """Vérifie le déséquilibre des classes"""
        class_counts = Counter(y)
        total_samples = len(y)
        min_class_ratio = min(class_counts.values()) / total_samples
        
        if min_class_ratio < threshold:
            return True
        return False
    
    def prepare_data(self, df, prediction_level='Classe'):
        """Préparation des données"""
        df['reponse_clean'] = df.apply(
            lambda row: self.preprocess_text(row['reponse'], row.get('langage', 'auto')), 
            axis=1
        )
        
        # Construire la structure hiérarchique
        self.build_hierarchy_structure(df)
        
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        target_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]
        
        for level in target_levels:
            if level not in self.label_encoders:
                le = LabelEncoder()
                df[f'{level}_encoded'] = le.fit_transform(df[level])
                self.label_encoders[level] = le
        
        return df
    
    # ----------------------------------------------------------------
    # ENTRAÎNEMENT AVEC PRISE EN COMPTE DES CORRECTIONS
    # ----------------------------------------------------------------

    def train_hierarchical_models(self, df, prediction_level='Classe'):
        """Compare Naive Bayes et Logistic Regression à chaque niveau hiérarchique"""
        # 🔁 Recalcul complet du TF-IDF (intègre les corrections)
        self.vectorizer = TfidfVectorizer(
            max_features=10000, ngram_range=(1, 3),
            lowercase=True, analyzer='word', min_df=1, max_df=0.95
        )

        X = self.vectorizer.fit_transform(df['reponse_clean'])
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        target_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]

        progress_bar = st.progress(0)
        status = st.empty()

        for idx, level in enumerate(target_levels):
            status.info(f"🔄 Entraînement au niveau: {level}")
            progress_bar.progress((idx + 1) / len(target_levels))

            encoded_col = f'{level}_encoded'
            if encoded_col not in df.columns:
                continue

            y = df[encoded_col]
            if len(y.unique()) < 2:
                st.warning(f"Pas assez de classes pour {level}")
                continue

            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.2, random_state=42, stratify=y
            )

#            if predictor.use_smote:
#                smote = SMOTE(random_state=42)
#                X_train, y_train = smote.fit_resample(X_train, y_train)

            best_model, best_name, best_score = None, None, 0.0
            for model_name, model in self.base_models.items():
                param_grid = {f"classifier__{k}": v for k, v in self.param_grids[model_name].items()}
                pipeline = Pipeline([('classifier', model)])
                grid = GridSearchCV(
                    pipeline, param_grid, cv=StratifiedKFold(3, shuffle=True, random_state=42),
                    scoring='f1_weighted', n_jobs=1
                )
                grid.fit(X_train, y_train)
                if grid.best_score_ > best_score:
                    best_model, best_name, best_score = grid.best_estimator_, model_name, grid.best_score_

            self.best_models[level] = {
                'name': best_name,
                'model': best_model,
                'score': best_score
            }
            st.success(f"✅ {level}")
            st.success(f"✅ {level}: {best_name} (F1={best_score:.3f})")


        progress_bar.empty()
        status.empty()

        # Initialisation des encodeurs de labels pour chaque niveau
#        label_encoders = {}
        
#        for level in target_levels:
#            if level not in label_encoders:
#                le = LabelEncoder()
#                df[f'{level}_encoded'] = le.fit_transform(df[level])
#                label_encoders[level] = le  # Conserver l'encodeur pour une utilisation ultérieure

        # Calculer et afficher le rapport de classification
        y_pred = best_model.predict(X_test)
        report = classification_report(y_test, y_pred)#, target_names=le.classes_)
        st.markdown("### Rapport de Classification")
        st.text(report)

    # ----------------------------------------------------------------
    # PRÉDICTION
    # ----------------------------------------------------------------

    def predict_hierarchy(self, text, prediction_level='Classe'):
        text_clean = self.preprocess_text(text)
        X = self.vectorizer.transform([text_clean])
        predictions, probabilities = {}, {}
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        target_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]

        for level in target_levels:
            if level not in self.best_models:
                predictions[level] = "Non disponible"
                probabilities[level] = 0.0
                continue

            model = self.best_models[level]['model']
            pred = model.predict(X)[0]
            proba = model.predict_proba(X)[0] if hasattr(model, 'predict_proba') else [1.0]
            label = self.label_encoders[level].inverse_transform([pred])[0]
            confidence = max(proba) if len(proba) > 0 else 0.0

            predictions[level] = label
            probabilities[level] = confidence

        return predictions, probabilities


def create_hierarchy_selector(df, predictor):
    """Crée les sélecteurs hiérarchiques avancés dans la sidebar"""
    st.sidebar.markdown("### 🎯 Sélection Hiérarchique")
    
    hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
    selections = {}
    
    # Section d'information sur la structure hiérarchique (simplifiée)
    with st.sidebar.expander("📊 Niveaux disponibles", expanded=False):
        for level in hierarchy_levels:
            unique_count = df[level].nunique()
            st.markdown(f"**{level}:** {unique_count} catégories")
    
    st.sidebar.markdown("---")
    
    # Sélecteurs hiérarchiques progressifs
    for i, level in enumerate(hierarchy_levels):
        if hasattr(predictor, 'hierarchy_structure') and predictor.hierarchy_structure:
            available_options = predictor.get_filtered_options(level, selections)
        else:
            if i == 0:
                available_options = sorted(df[level].unique())
            else:
                parent_level = hierarchy_levels[i-1]
                if parent_level in selections and selections[parent_level]:
                    filtered_df = df[df[parent_level] == selections[parent_level]]
                    available_options = sorted(filtered_df[level].unique())
                else:
                    available_options = []
        
        if available_options:
            options_with_all = ["-- Tous --"] + available_options
            
            selected = st.sidebar.selectbox(
                f"🔹 {level}",
                options=options_with_all,
                key=f"selector_{level}",
                help=f"Choisissez une catégorie pour {level}"
            )
            
            if selected != "-- Tous --":
                selections[level] = selected
                
                if hasattr(predictor, 'hierarchy_structure'):
                    is_unique, unique_class = predictor.get_unique_prediction_path(selections)
                    if is_unique:
                        st.sidebar.success("✅ Classification unique identifiée")
        else:
            if i > 0:
                st.sidebar.warning("⚠️ Aucune option disponible")
            break
    
    return selections

def process_batch_file(file_data, text_column):
    """Traite un fichier pour la prédiction par lot"""
    try:
        if file_data.name.endswith('.csv'):
            df = pd.read_csv(file_data)
        elif file_data.name.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_data)
        elif file_data.name.endswith('.txt'):
            content = file_data.read().decode('utf-8')
            texts = [line.strip() for line in content.split('\n') if line.strip()]
            df = pd.DataFrame({'texte': texts})
            text_column = 'texte'
        else:
            return None, None, "Format de fichier non supporté"
        
        if text_column not in df.columns:
            return None, None, f"Colonne '{text_column}' introuvable"
        
        texts = df[text_column].dropna().astype(str).tolist()
        return df, texts, None
        
    except Exception as e:
        return None, None, f"Erreur lors du traitement: {e}"

def create_batch_results_table(texts, predictions, probabilities, prediction_level):
    """Crée un tableau des résultats de prédiction par lot"""
    hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
    display_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]
    
    results_data = []
    for i, text in enumerate(texts):
        row = {'Texte': text[:100] + '...' if len(text) > 100 else text}
        
        for level in display_levels:
            if i < len(predictions) and level in predictions[i]:
                row[f'{level}'] = predictions[i][level]
                row[f'Confiance_{level}'] = f"{probabilities[i][level]:.2%}"
        
        results_data.append(row)
    
    return pd.DataFrame(results_data)

def create_prediction_analysis(predictions, probabilities, hierarchy_levels, selected_categories=None):
    """Crée une analyse détaillée des prédictions"""
    st.markdown("### 🎯 Analyse Détaillée des Prédictions")
    
    # Créer les données pour l'exportation
    comparison_data = []
    
    if selected_categories:
        st.markdown("#### 📊 Comparaison Prédiction vs Sélection")
        
        for level in hierarchy_levels:
            pred_value = predictions.get(level, "Non prédit")
            selected_value = selected_categories.get(level, "Non sélectionné")
            confidence = probabilities.get(level, 0.0)
            
            is_match = pred_value == selected_value
            match_icon = "✅" if is_match else "❌"
            
            comparison_data.append({
                "Niveau": level,
                "Prédiction": pred_value,
                "Sélection": selected_value,
                "Confiance": f"{confidence:.2%}",
                "Correspondance": f"{match_icon} {'Oui' if is_match else 'Non'}"
            })
        
        comparison_df = pd.DataFrame(comparison_data)
        st.dataframe(comparison_df, use_container_width=True)
        
        # Ajouter un bouton d'export
        if st.button("📥 Exporter les résultats"):
            try:
                from export_utils import export_to_excel
                output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "exports")
                filename = export_to_excel(comparison_data, output_dir)
                st.success(f"✅ Résultats exportés avec succès! Fichier: {filename}")
                full_path = os.path.join(output_dir, filename)
                with open(full_path, "rb") as file:
                    st.download_button(
                        label="📎 Télécharger le fichier Excel",
                        data=file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Erreur lors de l'exportation: {str(e)}")
        
        matches = sum(1 for level in hierarchy_levels 
                     if predictions.get(level) == selected_categories.get(level) and 
                        predictions.get(level) not in ["Non prédit", "Modèle non disponible"])
        total_compared = sum(1 for level in hierarchy_levels 
                           if predictions.get(level) not in ["Non prédit", "Modèle non disponible"] and 
                              selected_categories.get(level) != "Non sélectionné")
        
        if total_compared > 0:
            match_percentage = (matches / total_compared) * 100
            if match_percentage >= 80:
                st.success(f"🎯 Excellente correspondance: {match_percentage:.1f}% ({matches}/{total_compared})")
            elif match_percentage >= 60:
                st.warning(f"⚠️ Correspondance moyenne: {match_percentage:.1f}% ({matches}/{total_compared})")
            else:
                st.error(f"❌ Faible correspondance: {match_percentage:.1f}% ({matches}/{total_compared})")
    
    # Analyse de confiance par niveau
    st.markdown("#### 📈 Analyse de Confiance")
    confidence_data = []
    
    for level in hierarchy_levels:
        if level in predictions and level in probabilities:
            confidence = probabilities[level]
            if confidence > 0:
                confidence_data.append({
                    "Niveau": level,
                    "Confiance": confidence,
                    "Prédiction": predictions[level]
                })
    
    if confidence_data:
        conf_df = pd.DataFrame(confidence_data)
        
        fig_conf = go.Figure(data=[
            go.Bar(
                x=conf_df["Niveau"],
                y=conf_df["Confiance"],
                text=[f"{conf:.2%}" for conf in conf_df["Confiance"]],
                textposition='auto',
                marker_color=['#28a745' if conf > 0.8 else '#ffc107' if conf > 0.6 else '#dc3545' 
                             for conf in conf_df["Confiance"]]
            )
        ])
        
        fig_conf.update_layout(
            title="Niveaux de Confiance par Niveau Hiérarchique",
            xaxis_title="Niveaux",
            yaxis_title="Confiance",
            yaxis=dict(tickformat=".0%"),
            height=400,
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)'
        )
        
        st.plotly_chart(fig_conf, use_container_width=True)

#def update_model_with_corrections(predictor, df_prepared, corrections, prediction_level):
#    """Met à jour le modèle avec les corrections fournies"""
#    try:
#        new_data = []
#        for correction in corrections:
#            new_data.append({
#                'reponse': correction['texte'],
#                'reponse_clean': predictor.preprocess_text(correction['texte']),
#                'Grand poste': correction.get('Grand poste', ''),
#                'Section': correction.get('Section', ''),
#                'Groupe': correction.get('Groupe', ''),
#                'Classe': correction.get('Classe', '')
#            })
#        
#        correction_df = pd.DataFrame(new_data)
#        updated_df = pd.concat([df_prepared, correction_df], ignore_index=True)
#        
#        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
#        target_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]
#        
#        for level in target_levels:
#            if level in predictor.label_encoders:
#                le = predictor.label_encoders[level]
#                all_classes = sorted(set(list(le.classes_) + list(updated_df[level].unique())))
#                
#                from sklearn.preprocessing import LabelEncoder
#                new_le = LabelEncoder()
#                new_le.fit(all_classes)
#                predictor.label_encoders[level] = new_le
#                
#                updated_df[f'{level}_encoded'] = new_le.transform(updated_df[level])
#        
#        predictor.vectorizer = TfidfVectorizer(
#            max_features=5000,
#            ngram_range=(1, 3),
#            lowercase=True,
#            analyzer='word',
#            min_df=2,
#            max_df=0.95
#    )
#        predictor.train_hierarchical_models(updated_df, prediction_level)
#        
#        return updated_df, True
#        
#    except Exception as e:
#        st.error(f"Erreur lors de la mise à jour: {e}")
#        return df_prepared, False

def update_model_with_corrections(predictor, df_prepared, corrections, prediction_level):
    """Met à jour le modèle avec les corrections fournies."""
    try:
        new_data = []
        for correction in corrections:
            new_data.append({
                'reponse': correction['texte'],
                'reponse_clean': predictor.preprocess_text(correction['texte']),
                'Grand poste': correction.get('Grand poste', ''),
                'Section': correction.get('Section', ''),
                'Groupe': correction.get('Groupe', ''),
                'Classe': correction.get('Classe', '')
            })
        
        correction_df = pd.DataFrame(new_data)
        updated_df = pd.concat([df_prepared, correction_df], ignore_index=True)

        # Re-encoder toutes les classes dans le DataFrame mis à jour
        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
        for level in hierarchy_levels:
            if level in predictor.label_encoders:
                le = predictor.label_encoders[level]
                all_classes = sorted(set(list(le.classes_) + list(updated_df[level].unique())))
                
                new_le = LabelEncoder()
                new_le.fit(all_classes)
                updated_df[f'{level}_encoded'] = new_le.transform(updated_df[level])
                predictor.label_encoders[level] = new_le

        # Réentraîner le modèle sur les données mises à jour
        predictor.train_hierarchical_models(updated_df, prediction_level)
        
        # Sauvegarde du modèle
        data_hash = predictor.get_data_hash(updated_df)
        model_filename = predictor.get_model_filename(data_hash, prediction_level)
        predictor.save_models(model_filename)  # Sauvegarde des modèles

        st.success("🎯 Modèle mis à jour et sauvegardé avec succès !")
        return updated_df, True

    except Exception as e:
        st.error(f"Erreur lors de la mise à jour: {e}")
        return df_prepared, False

# Code à ajouter après la mise à jour du modèle
#for correction in valid_corrections:
#    test_text = correction['texte']
#    predictions, probabilities = predictor.predict_hierarchy(test_text, prediction_level)
#    st.success(f"Prédiction après mise à jour pour '{test_text}': {predictions} avec confiance {probabilities}")


@st.cache_data
def load_data(file_data=None):
    """Chargement des données avec cache"""
    if file_data is not None:
        try:
            df = pd.read_excel(file_data)
            return df
        except Exception as e:
            st.error(f"Erreur lors du chargement du fichier uploadé: {e}")
            return None
    else:
        try:
            data_path = os.path.join("Data", "Data.xlsx")
            if os.path.exists(data_path):
                df = pd.read_excel(data_path)
                return df
            else:
                st.warning(f"Fichier Data.xlsx introuvable dans le dossier Data/")
                return None
        except Exception as e:
            st.warning(f"Erreur lors du chargement de Data/Data.xlsx: {e}")
            return None



def main():
    # Initialisation du prédicteur
#    predictor = EnhancedHierarchicalPredictor()  

    # Afficher uniquement les logos en position fixe
    display_fixed_logos()

    # (No explicit active_tab needed) using 'mode_select' radio to control view

    # Charger le CSS et les éléments visuels
    load_css()
    
    # CSS FORCE OVERRIDE DIRECT - INJECTION MULTIPLE
    st.markdown("""
    <style>
        /* OVERRIDE ABSOLU - TOUS LES CHAMPS EN NOIR */
        .stSelectbox div[data-baseweb="select"] div,
        .stSelectbox div[role="combobox"],
        .stSelectbox div[data-testid="stSelectbox"] > div > div,
        .stSidebar .stSelectbox div[data-baseweb="select"] div,
        .stSidebar .stSelectbox div[role="combobox"],
        .stSidebar .stSelectbox div[data-testid="stSelectbox"] > div > div {
            background-color: #000000 !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 0 0 1px rgba(255, 255, 255, 0.1) !important;
        }
        
        /* TEXT AREAS ET INPUTS - FORCE NOIR - TOUS LES SELECTEURS POSSIBLES */
        .stTextArea textarea,
        .stTextInput input,
        .stSidebar .stTextArea textarea,
        .stSidebar .stTextInput input,
        textarea,
        input[type="text"],
        input[type="number"],
        div[data-testid="stTextArea"] textarea,
        div[data-testid="stTextInput"] input,
        [data-testid="textAreaInput"],
        [data-testid="textInput"] {
            background-color: #000000 !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 0 0 1px rgba(255, 255, 255, 0.1) !important;
        }
        
        /* PLACEHOLDER TEXT */
        textarea::placeholder,
        input::placeholder {
            color: rgba(255, 255, 255, 0.5) !important;
        }
        
        /* SELECTBOX DANS INTERFACE PRINCIPALE - FORCE NOIR */
        div[data-testid="stSelectbox"] div,
        div[data-testid="stSelectbox"] div div,
        .main .stSelectbox div,
        .main .stSelectbox div div,
        .main [data-baseweb="select"],
        .main [data-baseweb="select"] div {
            background-color: #000000 !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 0 0 1px rgba(255, 255, 255, 0.1) !important;
        }
        
        /* FOCUS STATES */
        .stSelectbox div[data-baseweb="select"] div:focus,
        .stTextArea textarea:focus,
        .stTextInput input:focus,
        div[data-testid="stSelectbox"] div:focus {
            box-shadow: 0 0 0 2px rgba(102, 126, 234, 0.5) !important;
            outline: none !important;
        }
        
        /* HOVER STATES */
        .stSelectbox div[data-baseweb="select"] div:hover,
        .stTextArea textarea:hover,
        .stTextInput input:hover,
        div[data-testid="stSelectbox"] div:hover {
            box-shadow: 0 0 0 1px rgba(255, 255, 255, 0.2) !important;
        }
        
        /* Style pour l'onglet actif */
        .stTabs [aria-selected="true"] {
            background-color: rgba(102, 126, 234, 0.2) !important;
            border-bottom-color: #667eea !important;
        }
        
        .stTabs [aria-selected="true"] p {
            color: #667eea !important;
            font-weight: bold !important;
        }
        
        /* FILE UPLOADER - FORCE NOIR PLUS SPECIFIQUE */
        .stFileUploader section,
        .stFileUploader div[data-testid="stFileUploaderDropzone"],
        .stSidebar .stFileUploader section,
        .stSidebar .stFileUploader div[data-testid="stFileUploaderDropzone"],
        div[data-testid="stFileUploader"] section,
        div[data-testid="stFileUploader"] div {
            background-color: #000000 !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 0 0 1px rgba(255,255,255,0.04) !important;
        }
        
        /* SIDEBAR BACKGROUND - FORCE NOIR */
        .css-1d391kg,
        .css-1y4p8pa,
        .css-12oz5g7,
        .stSidebar,
        section[data-testid="stSidebar"] {
            background-color: #1a1a1a !important;
        }
        
        /* SIDEBAR TEXT - FORCE BLANC */
        .stSidebar,
        .stSidebar *,
        .stSidebar p,
        .stSidebar label,
        .stSidebar span,
        .stSidebar div {
            color: #ffffff !important;
        }
        
        /* DATAFRAME - FORCE NOIR */
        .stDataFrame,
        .stDataFrame table,
        .stDataFrame tbody tr,
        .stDataFrame td,
        div[data-testid="dataframe"] table,
        div[data-testid="dataframe"] td {
            background-color: #000000 !important;
            color: #ffffff !important;
        }
        
        .stDataFrame th,
        div[data-testid="dataframe"] th {
            background-color: #1a1a1a !important;
            color: #ffffff !important;
        }
        
        /* FORCER TOUS LES DROPDOWNS */
        [role="listbox"],
        [role="option"],
        div[role="listbox"],
        div[role="option"] {
            background-color: #000000 !important;
            color: #ffffff !important;
        }
        
        /* SLIDER */
        .stSlider div {
            color: #ffffff !important;
        }
        
        /* EXPANDER */
        .stExpander,
        .stExpander div,
        div[data-testid="stExpander"] {
            background-color: #000000 !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 6px 18px rgba(0,0,0,0.45) inset, 0 1px 0 0 rgba(255,255,255,0.02) !important;
        }
    </style>
    """, unsafe_allow_html=True)

    set_background_image("structure.jpg")   ###

    # Corrections CSS ciblées pour assurer une lisibilité optimale tout en gardant un fond sombre
    st.markdown("""
    <style>
        /* Keep these main selector blocks dark but improve contrast */
        .hierarchy-selector, .batch-section, .update-section {
            background: linear-gradient(135deg, rgba(8,12,20,0.92) 0%, rgba(12,18,30,0.95) 100%) !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 6px 18px rgba(0,0,0,0.6) inset, 0 2px 8px rgba(0,0,0,0.4) !important;
        }

        /* Ensure textareas and inputs inside these dark sections are readable */
        .hierarchy-selector textarea,
        .batch-section textarea,
        .update-section textarea,
        .hierarchy-selector input,
        .batch-section input,
        .update-section input {
            background: rgba(0,0,0,0.75) !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: inset 0 1px 0 rgba(255,255,255,0.02) !important;
        }

        /* Keep sidebar expanders dark (affects "📊 Niveaux disponibles") */
        .stSidebar .stExpander {
            background: linear-gradient(180deg, rgba(12,16,22,0.95) 0%, rgba(8,10,14,0.95) 100%) !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.45) !important;
        }

        /* Main expanders keep readable dark style as well */
        .stExpander {
            background: linear-gradient(180deg, rgba(12,16,22,0.95) 0%, rgba(8,10,14,0.95) 100%) !important;
            color: #ffffff !important;
            border: none !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.45) !important;
        }

        .stSidebar .stExpander > div > div > div, .stExpander > div > div > div {
            color: #ffffff !important;
            font-weight: 600 !important;
        }
    </style>
    """, unsafe_allow_html=True)

    # En-tête principal
    # Style pour faire ressembler le radio 'mode_select' à des onglets
    st.markdown("""
    <style>
    /* Radio -> Tabs visual */
    div[role="radiogroup"]{ display:flex; gap:8px; margin-bottom:0.8rem; border-bottom:1px solid rgba(255,255,255,0.04); padding-bottom:6px; }
    div[role="radiogroup"] > div[role="radio"]{ background:transparent; color:rgba(255,255,255,0.85); padding:6px 14px; border-radius:10px 10px 0 0; cursor:pointer; transition:all .12s ease; }
    div[role="radiogroup"] > div[role="radio"]:hover{ background:rgba(255,255,255,0.03); color:#ffffff; }
    div[role="radiogroup"] > div[role="radio"][aria-checked="true"]{ background:linear-gradient(90deg,#667eea,#764ba2) !important; color:#ffffff !important; font-weight:700 !important; box-shadow:0 6px 18px rgba(0,0,0,0.5) !important; }
    /* Ensure the inner label/text uses full width */
    div[role="radiogroup"] > div[role="radio"] > label{ width:100%; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div class="main-header">
        <h1>Classifieur de l’activité selon la NACAM</h1>
        <p>Classification intelligente des activités exercées par un agent économique</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar principale
    st.sidebar.markdown("### **🎯 Classifieur de l'activité**")
    st.sidebar.markdown("---")
    st.sidebar.title("Configuration")
    
    # Upload du fichier de données
    st.sidebar.subheader("📊 Données")
    uploaded_file = st.sidebar.file_uploader(
        "Uploadez votre fichier Excel",
        type=['xlsx', 'xls'],
        help="Fichier contenant les données de nomenclature"
    )
    
    # Chargement des données
    if uploaded_file is not None:
        df = load_data(uploaded_file)
    else:
        df = load_data()
    
    if df is None:
        st.markdown("""
        <div class="prediction-result">
            <h3>Aucun fichier de données disponible</h3>
            <h4>Instructions:</h4>
            <ol>
                <li>Utilisez le bouton "Browse files" dans la sidebar</li>
                <li>Sélectionnez votre fichier Data.xlsx</li>
                <li>L'application se mettra à jour automatiquement</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)
        return
    
    # Validation des colonnes
    colonnes_requises = ['reponse', 'langage', 'Classe', 'Grand poste', 'Section', 'Groupe']
    colonnes_manquantes = [col for col in colonnes_requises if col not in df.columns]
    
    if colonnes_manquantes:
        st.sidebar.error(f"Colonnes manquantes: {', '.join(colonnes_manquantes)}")
        st.markdown(f"""
        <div class="prediction-result" style="border-color: #dc3545; background: rgba(220, 53, 69, 0.1);">
            <h3>Structure de fichier incorrecte</h3>
            <p><strong>Colonnes manquantes:</strong> {', '.join(colonnes_manquantes)}</p>
            <p><strong>Colonnes trouvées:</strong> {', '.join(df.columns.tolist())}</p>
        </div>
        """, unsafe_allow_html=True)
        return
    else:
        st.sidebar.success(f"Fichier valide - {len(df):,} entrées chargées")
    
    # Configuration des modèles
    st.sidebar.subheader("⚙️ Configuration")
    
    # Options avancées
    st.sidebar.markdown("#### Options d'optimisation")
    use_smote = st.sidebar.checkbox("Activer SMOTE (suréchantillonnage)", value=True, 
                                   help="Corrige les déséquilibres de classes")
    
    # Sélection du niveau de prédiction
    prediction_level = st.sidebar.selectbox(
        "Niveau de prédiction maximum",
        options=['Grand poste', 'Section', 'Groupe', 'Classe'],
        index=3,
        help="Niveau hiérarchique jusqu'auquel prédire"
    )
    
    # Information sur le modèle disponible (CPU optimisé)
    st.sidebar.markdown("#### 🤖 Modèle utilisé")
#    st.sidebar.info("**MultinomialNB** - Optimisé pour CPU")
    st.sidebar.markdown("*Paramètres optimisés automatiquement*")
    
    # Initialisation du prédicteur
    predictor = EnhancedHierarchicalPredictor()
    predictor.use_smote = use_smote
    
    # Préparer les données et construire la structure hiérarchique
    df_prepared = predictor.prepare_data(df, prediction_level)
    
    # Sélecteurs hiérarchiques dans la sidebar
    hierarchy_selections = create_hierarchy_selector(df_prepared, predictor)
    
    # Affichage des informations sur les sélections
    if hierarchy_selections:
        st.sidebar.markdown("### 📋 Sélections Actuelles")
        for level, selection in hierarchy_selections.items():
            st.sidebar.markdown(f"**{level}:** {selection}")
        
        # Vérifier si c'est un cas unique
        if hasattr(predictor, 'hierarchy_structure'):
            is_unique, unique_class = predictor.get_unique_prediction_path(hierarchy_selections)
            if is_unique:
                st.sidebar.success("✅ Classification unique identifiée")
    
    # Gestion de l'entraînement des modèles
    data_hash = predictor.get_data_hash(df_prepared)
    model_filename = predictor.get_model_filename(data_hash, prediction_level)
    
    st.sidebar.subheader("🧠 Gestion des Modèles")
    models_exist = predictor.models_exist(model_filename)
    
    if models_exist:
        st.sidebar.success("Modèles trouvés")
        use_saved = st.sidebar.checkbox("Utiliser modèles sauvegardés", value=True)
        force_retrain = st.sidebar.button("🔄 Re-entraîner")
    else:
        use_saved = False
        force_retrain = False
    
    # Cache pour les modèles
    cache_key = f"enhanced_predictor_{prediction_level}_{data_hash}_{use_smote}"
    
    # Entraînement ou chargement
    if models_exist and use_saved and not force_retrain:
        if cache_key not in st.session_state:
            with st.spinner("Chargement des modèles..."):
                try:
                    if predictor.load_models(model_filename):
                        st.session_state[cache_key] = {
                            'predictor': predictor,
                            'df_prepared': df_prepared
                        }
                        st.sidebar.success("Modèles chargés avec succès")
                except Exception as e:
                    st.sidebar.error(f"Erreur de chargement: {e}")
                    force_retrain = True
    
    if cache_key not in st.session_state or force_retrain:
        with st.spinner("Entraînement des modèles en cours..."):
#            st.info("🚀 Optimisation MultinomialNB en cours...")
            
            try:
                predictor.train_hierarchical_models(df_prepared, prediction_level)
                
                # Sauvegarder
                predictor.save_models(model_filename)
                
                st.session_state[cache_key] = {
                    'predictor': predictor,
                    'df_prepared': df_prepared
                }
                
            #    st.balloons()
                
            except Exception as e:
                st.error(f"Erreur lors de l'entraînement: {e}")
                return
    
    predictor = st.session_state[cache_key]['predictor']
    df_prepared = st.session_state[cache_key]['df_prepared']
    
    # Interface principale avec deux colonnes principales
    main_col1, main_col2 = st.columns([3, 1])
    
    with main_col1:
        st.header("🎯 Mode de Prédiction")
        
        # Mode selector (replace programmatic tabs to keep selection stable across reruns)
        mode = st.radio(
            "Onglet",
            ["Prédiction Manuelle", "Prédiction par Lot", "Mise à Jour Modèle"],
            index=0,
            key="mode_select",
            horizontal=True,
        )

        # Render sections conditionally based on selected mode
        if st.session_state.get("mode_select", "Prédiction Manuelle") == "Prédiction Manuelle":
            st.markdown("""
            <div class="hierarchy-selector">
                <h4>✍️ Prédiction Manuelle</h4>
                <p>Entrez un texte pour obtenir sa classification hiérarchique</p>
            </div>
            """, unsafe_allow_html=True)
            
            user_input = st.text_area(
                "Entrez votre texte à classifier:",
                placeholder="Ex: Culture de blé, Fabrication de meubles, Services bancaires...",
                height=100,
                key="manual_input"
            )
            
            col_predict, col_guided = st.columns(2)
            
            with col_predict:
                predict_button = st.button("🎯 Prédiction Libre", type="primary")
            
            with col_guided:
                guided_button = st.button("🧭 Prédiction Guidée", help="Utilise les sélections de la sidebar", type="primary")
            
            if predict_button and user_input:
                with st.spinner("Analyse en cours..."):
                    try:
                        predictions, probabilities = predictor.predict_hierarchy(user_input, prediction_level)
                        
                        st.markdown("### 📊 Résultats de la prédiction")
                        
                        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
                        display_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]
                        colors = ['#ff6b6b', '#4ecdc4', '#45b7d1', '#96ceb4']
                        
                        for i, level in enumerate(display_levels):
                            if level in predictions:
                                confidence = probabilities[level]
                                conf_class = "confidence-high" if confidence > 0.8 else "confidence-medium" if confidence > 0.6 else "confidence-low"
                                conf_text = "Élevée" if confidence > 0.8 else "Moyenne" if confidence > 0.6 else "Faible"
                                conf_icon = "✅" if confidence > 0.8 else "⚠️" if confidence > 0.6 else "❌"
                                
                                st.markdown(f"""
                                <div class="level-card">
                                    <h4 style="color: {colors[i]}; margin: 0 0 0.5rem 0;">
                                        {level}
                                    </h4>
                                    <p style="font-size: 1.1rem; margin: 0.5rem 0;">
                                        <strong>{predictions[level]}</strong>
                                    </p>
                                    <p style="margin: 0;">
                                        Confiance: <span class="{conf_class}">{confidence:.2%}</span> ({conf_text}) {conf_icon}
                                    </p>
                                </div>
                                """, unsafe_allow_html=True)
                        
                        # Graphique des probabilités
                        valid_levels = [level for level in display_levels 
                                      if predictions.get(level) not in ["Modèle non disponible", "Erreur de prédiction"]]
                        
                        if valid_levels:
                            fig = go.Figure(data=[
                                go.Bar(
                                    x=valid_levels,
                                    y=[probabilities[level] for level in valid_levels],
                                    marker_color=[colors[hierarchy_levels.index(level)] for level in valid_levels],
                                    text=[f"{probabilities[level]:.2%}" for level in valid_levels],
                                    textposition='auto'
                                )
                            ])
                            
                            fig.update_layout(
                                title=f"Niveaux de confiance - Prédiction {prediction_level}",
                                xaxis_title="Niveaux hiérarchiques",
                                yaxis_title="Niveau de confiance",
                                yaxis=dict(tickformat=".0%"),
                                height=400,
                                plot_bgcolor='rgba(0,0,0,0)',
                                paper_bgcolor='rgba(0,0,0,0)'
                            )
                            
                            st.plotly_chart(fig, use_container_width=True)
                        
                    except Exception as e:
                        st.error(f"Erreur lors de la prédiction: {e}")
            
            if guided_button and user_input:
                with st.spinner("Analyse guidée en cours..."):
                    try:
                        predictions, probabilities = predictor.predict_hierarchy(user_input, prediction_level)
                        
                        # Définir les niveaux hiérarchiques pour l'analyse
                        hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
                        target_hierarchy_levels = hierarchy_levels[:hierarchy_levels.index(prediction_level) + 1]
                        
                        # Créer l'analyse comparative avec les sélections
                        create_prediction_analysis(predictions, probabilities, target_hierarchy_levels, hierarchy_selections)
                        
                    except Exception as e:
                        st.error(f"Erreur lors de la prédiction guidée: {e}")
        
    if st.session_state.get("mode_select", "Prédiction Manuelle") == "Prédiction par Lot":
            st.markdown("""
            <div class="batch-section">
                <h4>📂 Prédiction par Lot</h4>
                <p>Uploadez un fichier contenant plusieurs textes à classifier</p>
            </div>
            """, unsafe_allow_html=True)
            
            # Gérer le changement d'onglet quand un fichier est uploadé
            def on_file_upload():
                # When a batch file is uploaded, set the mode radio to the batch view
                st.session_state["mode_select"] = "Prédiction par Lot"

            uploaded_batch_file = st.file_uploader(
                "Choisissez votre fichier",
                type=['csv', 'xlsx', 'xls', 'txt'],
                help="Formats supportés: CSV, Excel, TXT",
                key="batch_file",
                on_change=on_file_upload
            )
            
            if uploaded_batch_file is not None:
                col_config1, col_config2 = st.columns(2)
                
                with col_config1:
                    if uploaded_batch_file.name.endswith('.txt'):
                        text_column = "texte"
                        st.info("Fichier TXT détecté - une ligne = un texte")
                    else:
                        try:
                            if uploaded_batch_file.name.endswith('.csv'):
                                preview_df = pd.read_csv(uploaded_batch_file, nrows=3)
                            else:
                                preview_df = pd.read_excel(uploaded_batch_file, nrows=3)
                            
                            text_column = st.selectbox(
                                "Colonne contenant les textes:",
                                options=preview_df.columns.tolist(),
                                key="text_column_select"
                            )
                            
                            st.markdown("**Aperçu du fichier:**")
                            st.dataframe(preview_df, use_container_width=True)
                            
                        except Exception as e:
                            st.error(f"Erreur lors de la lecture du fichier: {e}")
                            text_column = None
                
                with col_config2:
                    batch_size = st.slider(
                        "Taille du lot:",
                        min_value=10,
                        max_value=500,
                        value=100
                    )
                
                if st.button("📂 Traiter le lot", type="primary") and text_column:
                    with st.spinner("Traitement du fichier en cours..."):
                        df_batch, texts_batch, error = process_batch_file(uploaded_batch_file, text_column)
                        
                        if error:
                            st.error(f"Erreur: {error}")
                        elif texts_batch:
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            total_texts = len(texts_batch)
                            all_predictions = []
                            all_probabilities = []
                            
                            for i in range(0, total_texts, batch_size):
                                batch_texts = texts_batch[i:i + batch_size]
                                status_text.text(f"Traitement: {i + 1}-{min(i + batch_size, total_texts)} sur {total_texts}")
                                
                                try:
                                    for text in batch_texts:
                                        pred, prob = predictor.predict_hierarchy(text, prediction_level)
                                        all_predictions.append(pred)
                                        all_probabilities.append(prob)
                                except Exception as e:
                                    st.error(f"Erreur lors du traitement: {e}")
                                    break
                                
                                progress_bar.progress(min(1.0, (i + batch_size) / total_texts))
                            
                            progress_bar.empty()
                            status_text.empty()
                            
                            if all_predictions:
                                st.success(f"Traitement terminé! {len(all_predictions)} textes classifiés")
                                
                                results_df = create_batch_results_table(
                                    texts_batch, all_predictions, all_probabilities, prediction_level
                                )
                                
                                st.markdown("### 📊 Résultats de la classification par lot")
                                st.dataframe(results_df, use_container_width=True, height=400)
                                
                                # Options d'exportation
                                with st.expander("📥 Options d'exportation", expanded=True):
                                    export_col1, export_col2 = st.columns([1, 2])
                                    
                                    with export_col1:
                                        export_format = st.selectbox(
                                            "Format d'export",
                                            options=["Excel (.xlsx)", "CSV (.csv)"],
                                            index=0,
                                            help="Choisissez le format du fichier d'export"
                                        )
                                    
                                    with export_col2:
                                        include_confidence = st.checkbox("✓ Niveaux de confiance", 
                                            value=True,
                                            help="Inclure les scores de confiance pour chaque prédiction")
                                        include_source = st.checkbox("✓ Données sources", 
                                            value=True,
                                            help="Inclure les textes originaux et prétraités")
                                        include_metadata = st.checkbox("✓ Métadonnées", 
                                            value=True,
                                            help="Inclure les informations sur le modèle et la date de prédiction")

                                    # Bouton d'export unique qui déclenchera l'action
                                    export_clicked = st.button("💾 Exporter les résultats", use_container_width=True)

                                    # Ne procéder à l'export que si le bouton est cliqué
                                    if export_clicked:
                                        # Préparation des données pour l'export
                                        export_df = results_df.copy()
                                
                                # Préparation des données d'export
                                export_df = results_df.copy()
                                
                                # Ajout minimal des métadonnées (date/heure) si demandé
                                metadata_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                if include_metadata:
                                    # Ajouter une colonne Date_prediction pour traçabilité
                                    export_df['Date_prediction'] = metadata_date
                                
                                # Suppression des colonnes si non demandées
                                if not include_confidence:
                                    confidence_cols = [col for col in export_df.columns if 'confiance' in col.lower() or 'probabilité' in col.lower()]
                                    export_df = export_df.drop(columns=confidence_cols, errors='ignore')
                                
                                if not include_source:
                                    source_cols = ['texte_original', 'texte_pretraite']
                                    export_df = export_df.drop(columns=[col for col in source_cols if col in export_df.columns])
                                
                                # Initialiser les variables de session pour les options d'export si elles n'existent pas
                                if 'export_format' not in st.session_state:
                                    st.session_state.export_format = export_format
                                if 'include_metadata' not in st.session_state:
                                    st.session_state.include_metadata = include_metadata

                                # Export selon le format choisi
                                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                if st.session_state.export_format == "Excel (.xlsx)":
                                    output = BytesIO()
                                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                        export_df.to_excel(writer, index=False, sheet_name='Prédictions')

                                        # Feuille de métadonnées (date/heure uniquement)
                                        if include_metadata:
                                            pd.DataFrame([{'Date_prediction': metadata_date}]).to_excel(
                                                writer,
                                                sheet_name='Métadonnées',
                                                index=False
                                            )

                                        # Appliquer le style conditionnel aux colonnes de confiance
                                        try:
                                            workbook = writer.book
                                            worksheet = writer.sheets['Prédictions']

                                            # Définir les fills
                                            high_fill = PatternFill(start_color="FF28A745", end_color="FF28A745", fill_type="solid")
                                            medium_fill = PatternFill(start_color="FFFFC107", end_color="FFFFC107", fill_type="solid")
                                            low_fill = PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid")

                                            # Parcourir les colonnes pour trouver celles de confiance
                                            for idx, col in enumerate(export_df.columns, start=1):
                                                if 'confiance' in col.lower() or 'probabilité' in col.lower():
                                                    col_letter = get_column_letter(idx)
                                                    for row_idx in range(2, len(export_df) + 2):
                                                        cell = worksheet[f"{col_letter}{row_idx}"]
                                                        val = cell.value
                                                        v = None
                                                        # Gérer les formats '12.34%' ou valeurs numériques
                                                        try:
                                                            if isinstance(val, str) and val.strip().endswith('%'):
                                                                v = float(val.strip().replace('%', '')) / 100.0
                                                            else:
                                                                v = float(val)
                                                        except Exception:
                                                            v = None

                                                        if v is not None:
                                                            if v > 0.8:
                                                                cell.fill = high_fill
                                                            elif v > 0.6:
                                                                cell.fill = medium_fill
                                                            else:
                                                                cell.fill = low_fill
                                        except Exception as e:
                                            # Si le styling échoue, on ne bloque pas l'export
                                            import traceback
                                            traceback.print_exc()

                                    excel_data = output.getvalue()
                                    st.download_button(
                                        label="💾 Télécharger les résultats (Excel)",
                                        data=excel_data,
                                        file_name=f"predictions_hierarchiques_{timestamp}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                                else:
                                    csv_data = export_df.to_csv(index=False)
                                    st.download_button(
                                        label="💾 Télécharger les résultats (CSV)",
                                        data=csv_data,
                                        file_name=f"predictions_hierarchiques_{timestamp}.csv",
                                        mime="text/csv"
                                    )
                        else:
                            st.warning("Aucun texte trouvé dans le fichier")
        

    if st.session_state.get("mode_select", "Prédiction Manuelle") == "Mise à Jour Modèle":
        st.markdown("""
        <div class="update-section">
            <h4>🔄 Mise à Jour du Modèle</h4>
            <p>Corrigez les prédictions incorrectes pour améliorer le modèle</p>
        </div>
        """, unsafe_allow_html=True)
    
        st.markdown("#### Ajouter une correction")
    
        # ✅ Initialiser les clés hiérarchiques vides dans la session
        for key in ["correct_gp", "correct_section", "correct_groupe", "correct_classe"]:
            st.session_state.setdefault(key, "")
    
        col_text, col_correct = st.columns([2, 1])
    
        # 🧾 Zone de texte pour le texte mal prédit
        with col_text:
            correction_text = st.text_area(
                "Texte mal prédit:",
                placeholder="Entrez le texte qui a été mal classifié...",
                height=80,
                key="correction_text"
            )
    
        # 🧩 Sélecteurs hiérarchiques
        with col_correct:
            st.markdown("**Classifications correctes:**")
    
            correct_grand_poste = st.selectbox(
                "Grand poste correct:",
                options=[""] + sorted(df_prepared['Grand poste'].unique()),
                key="correct_gp"
            )
    
            filtered_sections = (
                df_prepared[df_prepared['Grand poste'] == correct_grand_poste]['Section'].unique()
                if correct_grand_poste else []
            )
            correct_section = st.selectbox(
                "Section correcte:",
                options=[""] + sorted(filtered_sections),
                key="correct_section"
            )
    
            filtered_groups = (
                df_prepared[
                    (df_prepared['Grand poste'] == correct_grand_poste) &
                    (df_prepared['Section'] == correct_section)
                ]['Groupe'].unique()
                if correct_section else []
            )
            correct_groupe = st.selectbox(
                "Groupe correct:",
                options=[""] + sorted(filtered_groups),
                key="correct_groupe"
            )
    
            filtered_classes = (
                df_prepared[
                    (df_prepared['Grand poste'] == correct_grand_poste) &
                    (df_prepared['Section'] == correct_section) &
                    (df_prepared['Groupe'] == correct_groupe)
                ]['Classe'].unique()
                if correct_groupe else []
            )
            correct_classe = st.selectbox(
                "Classe correcte:",
                options=[""] + sorted(filtered_classes),
                key="correct_classe"
            )
    
        # ---------- Initialiser la liste des corrections ----------
        if "corrections" not in st.session_state:
            st.session_state.corrections = []
    
        # ---------- Colonnes d'action ----------
        col_add, col_train = st.columns(2)
    
        with col_add:
            if st.button("➕ Ajouter Correction"):
                correction_text = st.session_state.get("correction_text", "").strip()
                correct_grand_poste = st.session_state.get("correct_gp", "").strip()
                correct_section = st.session_state.get("correct_section", "").strip()
                correct_groupe = st.session_state.get("correct_groupe", "").strip()
                correct_classe = st.session_state.get("correct_classe", "").strip()
    
                if not correction_text:
                    st.error("❗ Le texte de la correction est vide.")
                elif not all([correct_grand_poste, correct_section, correct_groupe, correct_classe]):
                    st.error("❗ Veuillez sélectionner tous les niveaux hiérarchiques (Grand poste / Section / Groupe / Classe).")
                else:
                    correction = {
                        "texte": correction_text,
                        "Grand poste": correct_grand_poste,
                        "Section": correct_section,
                        "Groupe": correct_groupe,
                        "Classe": correct_classe,
                    }
                    st.session_state.corrections.append(correction)
                    st.success("✅ Correction ajoutée avec succès !")
    
                    # 🔄 Réinitialiser les champs pour faciliter la saisie suivante
                    #st.session_state.correction_text = ""
                    #st.session_state.correct_gp = ""
                    #st.session_state.correct_section = ""
                    #st.session_state.correct_groupe = ""
                    #st.session_state.correct_classe = ""
    

        with col_train:
            if st.button("🔄 Mettre à jour le modèle", type="primary"):
                corrections = st.session_state.get("corrections", [])
                if not corrections:
                    st.info("ℹ️ Aucune correction en attente.")
                else:
                    valid_corrections = []  # INITIALISATION ICI
                    invalid_entries = 0

                    # Validation des champs de chaque correction
                    for corr in corrections:
                        for k in ["Grand poste", "Section", "Groupe", "Classe", "texte"]:
                            corr[k] = str(corr.get(k, "")).strip()
                        if all(corr.get(k) for k in ["Grand poste", "Section", "Groupe", "Classe", "texte"]):
                            valid_corrections.append(corr)
                        else:
                            invalid_entries += 1

                    if invalid_entries > 0:
                        st.warning(f"⚠️ {invalid_entries} correction(s) ignorée(s) car incomplètes.")
                    if not valid_corrections:
                        st.error("🚫 Aucune correction valide à intégrer. Vérifiez les champs hiérarchiques et le texte.")
                    else:
                        with st.spinner("Mise à jour du modèle en cours..."):
                            df_prepared, updated = update_model_with_corrections(
                                predictor, df_prepared, valid_corrections, prediction_level
                            )
                        if updated:
                            st.success("🎯 Modèle mis à jour avec succès avec les corrections valides !")
                            # Optionnel : vider les corrections après mise à jour
                            st.session_state.corrections = []

                            # Validation des nouvelles prédictions
#                            for correction in valid_corrections:
#                                test_text = correction['texte']
#                                predictions, probabilities = predictor.predict_hierarchy(test_text, prediction_level)
#                                st.success(f"Prédiction après mise à jour pour '{test_text}': {predictions} avec confiance {probabilities}")


                            st.markdown("### 🔍 Analyse des Prédictions après Mise à Jour")
                            for correction in valid_corrections:
                                test_text = correction['texte']
                                actual_class = correction['Classe']
                                predictions, probabilities = predictor.predict_hierarchy(test_text, prediction_level)
                                
                                st.write(f"Texte: {test_text}")
                                st.write(f"Prédiction: {predictions} avec confiance {probabilities}")
                                st.write(f"Classe réelle: {actual_class}")

#        with col_train:
#            if st.button("🔄 Mettre à jour le modèle", type="primary"):
#                corrections = st.session_state.get("corrections", [])
#                if not corrections:
#                    st.info("ℹ️ Aucune correction en attente.")
#                else:
#                    valid_corrections = []
#                    invalid_entries = 0
#    
#                    # Validation des champs de chaque correction
#                    for corr in corrections:
#                        for k in ["Grand poste", "Section", "Groupe", "Classe", "texte"]:
#                            corr[k] = str(corr.get(k, "")).strip()
#                        if all(corr.get(k) for k in ["Grand poste", "Section", "Groupe", "Classe", "texte"]):
#                            valid_corrections.append(corr)
#                        else:
#                            invalid_entries += 1
#    
#                    if invalid_entries > 0:
#                        st.warning(f"⚠️ {invalid_entries} correction(s) ignorée(s) car incomplètes.")
#                    if not valid_corrections:
#                        st.error("🚫 Aucune correction valide à intégrer. Vérifie les champs hiérarchiques et le texte.")
#                    else:
#                        with st.spinner("Mise à jour du modèle en cours..."):
#                            df_prepared, updated = update_model_with_corrections(
#                                predictor, df_prepared, valid_corrections, prediction_level
#                            )
#                        if updated:
#                            st.success("🎯 Modèle mis à jour avec succès avec les corrections valides !")
                            # Optionnel : vider les corrections après mise à jour
                            # st.session_state.corrections = []
    
        # ---------- Affichage des corrections ----------
        if st.session_state.corrections:
            st.markdown("#### Corrections en attente:")
            for i, correction in enumerate(st.session_state.corrections):
                with st.expander(f"Correction {i+1}: {correction['texte'][:50]}..."):
                    st.json(correction)
                    if st.button(f"🗑️ Supprimer", key=f"delete_{i}"):
                        st.session_state.corrections.pop(i)
                        st.experimental_rerun()
    

    with main_col2:
        # N'afficher les informations que si nous ne sommes pas dans le mode 'Prédiction par Lot'
        if st.session_state.get("mode_select", "Prédiction Manuelle") not in ["Prédiction par Lot","Mise à Jour Modèle"]:
            st.header("ℹ️ Informations")
            
            # Informations sur les niveaux uniquement
            hierarchy_levels = ['Grand poste', 'Section', 'Groupe', 'Classe']
            for level in hierarchy_levels:
                unique_count = df_prepared[level].nunique()
                st.markdown(f"""
                <div class="hierarchy-info">
                    <strong>{level}</strong><br>
                    📊 {unique_count} catégories
                </div>
                """, unsafe_allow_html=True)
        
        # Informations sur le modèle
        
        # Sélections actuelles
        if hierarchy_selections:
            st.markdown("### 🎯 Sélections Actuelles")
            for level, selection in hierarchy_selections.items():
                st.markdown(f"**{level}:** {selection}")
            
            # Vérifier si c'est un cas unique
            if hasattr(predictor, 'hierarchy_structure'):
                is_unique, unique_class = predictor.get_unique_prediction_path(hierarchy_selections)
                if is_unique:
                    st.success("✅ Classification unique identifiée")
        
        # Corrections en attente
        if 'corrections' in st.session_state and st.session_state.corrections:
            st.markdown("### 🔄 Corrections en Attente")
            st.info(f"{len(st.session_state.corrections)} correction(s) en attente")
        
        # Aide (ne pas afficher dans le mode de prédiction par lot)
        if st.session_state.get("mode_select", "Prédiction Manuelle") != "Prédiction par Lot":
            with st.expander("❓ Aide", expanded=False):
                st.markdown("""
                **Prédiction Manuelle:**
                - Libre: Prédiction sans contrainte
                - Guidée: Utilise vos sélections
                
                **Prédiction par Lot:**
                - Uploadez CSV, Excel ou TXT
                - Traitez plusieurs textes
                
                **Mise à Jour:**
                - Corrigez les erreurs de prédiction
                - Améliorez le modèle en continu
                """)
    
    # Footer simplifié
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; padding: 1rem;">
        <p style="color: rgba(255,255,255,0.8); font-size: 0.9rem;">
            🎯 Prédicteur Hiérarchique MultinomialNB - Optimisé CPU
        </p>
        <p style="color: rgba(255,255,255,0.6); font-size: 0.8rem;">
            Fonctionnalités: Prédiction manuelle • Traitement par lot • Mise à jour continue
        </p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":

    main()

