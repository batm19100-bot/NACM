import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def export_to_excel(predictions_data, output_path):
    """
    Exporte les prédictions vers un fichier Excel avec mise en forme conditionnelle.
    
    Args:
        predictions_data (list): Liste de dictionnaires contenant les prédictions
        output_path (str): Chemin où sauvegarder le fichier Excel
    """
    try:
        # Créer un DataFrame avec les prédictions
        df = pd.DataFrame(predictions_data)
        
        # Créer le nom du fichier avec horodatage
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"predictions_{timestamp}.xlsx"
        full_path = os.path.join(output_path, filename)
        
        # Assurer que le dossier existe
        os.makedirs(output_path, exist_ok=True)
        
        # Exporter vers Excel
        df.to_excel(full_path, index=False, engine='openpyxl')
        
        # Appliquer la mise en forme conditionnelle
        from openpyxl import load_workbook
        wb = load_workbook(full_path)
        ws = wb.active
        
        # Définir les couleurs pour la mise en forme conditionnelle
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        
        # Trouver la colonne de confiance
        confidence_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Confiance':
                confidence_col = col
                break
        
        if confidence_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=confidence_col)
                try:
                    # Convertir la valeur de confiance en nombre
                    confidence = float(str(cell.value).strip('%')) / 100
                    
                    # Appliquer la couleur selon le niveau de confiance
                    if confidence > 0.8:
                        cell.fill = green_fill
                    elif confidence > 0.6:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
                except (ValueError, TypeError):
                    continue
        
        # Ajuster la largeur des colonnes
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Sauvegarder le fichier avec la mise en forme
        wb.save(full_path)
        return filename
        
    except Exception as e:
        raise Exception(f"Erreur lors de l'exportation Excel: {str(e)}")